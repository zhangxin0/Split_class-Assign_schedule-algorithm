"""
SA排课算法
"""
import random
import openpyxl
import copy
import gc
import time
import functools
import sys
from . import data_source as data_source


class AssignSchedule:
    """

     01: initialize

        初始化用户输入参数：
        course_hours： 记录各个科目需要排课的总课时数
        free_course_hours: 记录安排完固定课程后，非连堂课需要的总课时数
        linked_course_hours_classes: 各个班级所需排连堂课课时数
        courses: 所有的科目
        flow_classes_time: 走班课时间
        manual_assign_time: 手排课程时间
        linked_course_times: 连堂课可排时间
        linked_course_hours: 连堂课的标准所需课时数
        class_num: 班级对应的数字号, eg {'1班'：1, ...}
        class_amount: 班级数
        subj_no_assign: 固定科目不排课
        classes_name: 班级名称
        static_week: 周上课天数
        static_section: 日课时数
        teacher_no_assign: 固定老师不排课
        conflict_course_list: 记录所有不可交换点

        重要参数：
        classes_subj: 记录classes_subj【班级】【科目】是否已排
        day_classes: 记录某课程在某天是否已排 day_classes[cla][day][subj] == 1
        schedules: 记录所有班级的课表 schedules[cla][day][section] = 'subj'
        linked_course_time_all_classes: 班级可排连堂课容器 —— 记录所有班级的连堂课可排时间 初始化
            = linked_course_times - manual_assign_time - flow_classes_time
        free_position_all_classes： 班级可排位置容器 —— 记录固定课程、连堂课排完后，所有班级剩余课程（以下统称：free 课程）的可排时间
        already_assigned_free_positions: 已排free课程的位置， 用于与冲突交换
        already_assigned_linked_positions：已排连堂课位置， 用于冲突点交换
    """

    def __init__(self, data):
        self.data = data
        # class_hours to record how many classes left
        self.course_hours = self.data.COURSE_TOTAL
        self.free_course_hours = []
        self.linked_course_hours_classes = []
        self.class_course_teacher_info = self.data.CLASSES_COURSE_TEACHER
        self.courses = self.data.COURSE
        self.classes = self.data.CLASSES
        self.flow_classes_time = self.data.FLOW_CLASS_TIME
        self.not_assign_time = self.data.NOT_ASSIGN_TIME
        self.manual_assign_time = self.data.PART_SCHEDULE
        self.linked_course_times = self.data.LINK_COURSE_TIMES
        self.linked_course_hours = self.data.COURSE_LINKED_HOURS
        self.class_num = self.data.CLASSES_NUM
        self.class_amount = self.data.CLASS_NUM
        self.subj_no_assign = self.data.NOT_ASSIGN_COURSE
        self.classes_name = self.data.CLASSES
        self.static_week = self.data.WEEK
        self.static_section = self.data.SECTION
        self.teacher_no_assign = self.data.NOT_ASSIGN_TEACHER
        self.conflict_course_list = []
        self.locked_cla_course = self.data.LOCKED_CLA_COURSE
        self.extra_assign_time = self.data.EXTRA_ASSIGN_COURSE

        self.link_course_count = self.data.LINK_COURSE_COUNT_PEER_DAY
        # define classes_subj : whether class has finished assignment subj
        # [[0, 0, 0, 0, 0, 0], ...[...]]
        self.classes_subj = []
        for i in range(self.data.CLASS_NUM):
            cla = {}
            for subj in self.class_course_teacher_info[i + 1]:
                cla[subj] = 0
            self.classes_subj.append(cla)
        self.day_classes = []
        # self.classes_linked_subj = []
        for i in range(self.class_amount):
            self.linked_course_hours_classes.append(copy.deepcopy(self.linked_course_hours))
        for i in range(self.data.CLASS_NUM):
            week = []
            for d in range(5):
                day = {}
                for subj in self.class_course_teacher_info[i + 1]:
                    day[subj] = 0
                week.append(day)
            self.day_classes.append(week)
        self.schedules = []
        for i in range(self.data.CLASS_NUM):
            week = []
            for d in range(5):
                day = []
                for j in range(self.data.SECTION):
                    day.append(None)
                week.append(day)
            self.schedules.append(week)
        # define linked_course_times for each class:
        # 初始化的时候移除了已排课程
        self.linked_course_time_all_classes = []
        for m in range(self.class_amount):
            linked_course_time_a_classes = copy.deepcopy(self.linked_course_times)
            # linked_course_time_a_classes - fixed schedule
            for index in range(len(linked_course_time_a_classes) - 1, -1, -1):
                link = linked_course_time_a_classes[index]
                flag = False
                for assigned in self.manual_assign_time:
                    # bug 11: index with num
                    if self.class_num[assigned['class']] - 1 == m:
                        # bug 10: 连堂课排两节，都不能与已排课程冲突
                        if (assigned['week'] == link['week'] and assigned['section'] == link['section']) or \
                                (assigned['week'] == link['week'] and assigned['section'] == link['section'] + 1):
                            linked_course_time_a_classes.pop(index)
                            flag = True
                            break
                if flag:
                    continue
                for flow in self.flow_classes_time:
                    if flow['week'] == link['week'] and flow['section'] == link['section'] or \
                            flow['week'] == link['week'] and flow['section'] == link['section'] + 1:
                        linked_course_time_a_classes.pop(index)
                        break
            self.linked_course_time_all_classes.append(linked_course_time_a_classes)
        # define free positions for each class:
        self.free_position_all_classes = []
        # define free already assigned positions:
        self.already_assigned_free_positions = []
        # define linked already assigned positions:
        self.already_assigned_linked_positions = []
        for cla in range(self.class_amount):
            self.already_assigned_free_positions.append([])
            self.already_assigned_linked_positions.append([])

    """
    02 - 01: Assign locked courses
    
    02 - 02: Assign fixed classes:
        a. 安排固定课程到课表中，包括：
           走班课程，手排课
        b. 手排课安排完成后，更新连堂课剩余的课时数，用来安排后续的连堂课
    """

    def assign_fixed(self):
        # assign flow classes:
        for time in self.flow_classes_time:
            for cla in range(len(self.classes)):
                day = time['week'] - 1
                section = time['section'] - 1
                self.schedules[cla][day][section] = self.data.FLOW_SIGN
        # assign manual assigned classes:
        for manual_time in self.manual_assign_time:
            class_name = manual_time['class']
            # bug3: cla_num get real cla -1 to index
            cla = self.class_num[class_name] - 1
            day = manual_time['week'] - 1
            section = manual_time['section'] - 1
            course = manual_time['course']
            self.schedules[cla][day][section] = course
            # update status of day and subject:
            self.day_classes[cla][day][course] = 1
            # 这里，如果该课程只排了一部分，比如体育课只排1节，不能把该科目去掉不排
            # self.classes_subj[cla][course] = 1
        # update classes_linked_course_hours:
        for i in range(len(self.manual_assign_time)):
            cla = self.class_num[self.manual_assign_time[i]['class']] - 1
            course = self.manual_assign_time[i]['course']
            if 'link' in self.manual_assign_time[i]:
                if self.manual_assign_time[i]['link']:
                    self.linked_course_hours_classes[cla][course] -= 1
        # 额外安排的课程linked_hours = 0
        for cla in self.extra_assign_time:
            cla_num = self.class_num[cla]
            for unit in self.extra_assign_time[cla]:
                course = unit['course']
                self.linked_course_hours_classes[cla_num - 1][course] = 0
        # assign not assign classes:
        for time in self.not_assign_time:
            for cla in range(len(self.classes)):
                day = time['week'] - 1
                section = time['section'] - 1
                self.schedules[cla][day][section] = self.data.NOT_ASSIGN_SIGN

    def assign_locked_course_linked(self):
        self.update_linked_positions()
        for cla in range(self.class_amount):
            for subj in self.class_course_teacher_info[cla + 1]:
                key = str(self.classes[cla]) + '-' + subj
                related_cla = self.get_related_class(cla, subj)
                if key in self.locked_cla_course:
                    cla_lock = self.locked_cla_course[key]['class']
                    cla_lock = self.class_num[cla_lock] - 1
                    subj_lock = self.locked_cla_course[key]['subject']
                    related_cla_lock = self.get_related_class(cla_lock, subj_lock)
                    # Assign subj for cla meanwhile assign subj_lock for cla_lock:
                    linked_hours = self.linked_course_hours_classes[cla][subj]
                    linked_hours_locked = self.linked_course_hours_classes[cla_lock][subj_lock]
                    hours = min(linked_hours, linked_hours_locked)
                    cla_time_linked = self.linked_course_time_all_classes[cla]
                    subj_time_linked = copy.deepcopy(cla_time_linked)
                    cla_time_locked_linked = self.linked_course_time_all_classes[cla_lock]
                    subj_time_locked_linked = copy.deepcopy(cla_time_locked_linked)
                    # 约束条件增加位置1：
                    # linked_course subj_time - subj_no_assign
                    self.remove_subj_no_assign(subj, subj_time_linked)
                    # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
                    self.remove_already_related_conflict_course_position_linked(subj, related_cla,
                                                                                subj_time_linked)
                    # linked_course subj_time - teacher_no_assign
                    self.remove_teacher_no_assign(subj, cla, subj_time_linked)
                    # 移除所有已经排了2天连堂课的天数：
                    self.remove_already_assigned_two_linked_classes(cla, subj_time_linked)
                    # Also remove subj_time_linked_locked:
                    # -----------------------------------------------------------------------------------
                    # linked_course subj_time - subj_no_assign
                    self.remove_subj_no_assign(subj_lock, subj_time_locked_linked)
                    # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
                    self.remove_already_related_conflict_course_position_linked(subj_lock, related_cla_lock,
                                                                                subj_time_locked_linked)
                    # linked_course subj_time - teacher_no_assign
                    self.remove_teacher_no_assign(subj_lock, cla_lock, subj_time_locked_linked)
                    # 移除所有已经排了2天连堂课的天数：
                    self.remove_already_assigned_two_linked_classes(cla_lock, subj_time_locked_linked)
                    # Two locked subj 取交集：
                    both_ok_position = []
                    for grid in subj_time_linked:
                        if grid in subj_time_locked_linked:
                            both_ok_position.append(grid)
                    # BOTH_OK_POSITION 移除最后一个位置
                    for j in range(len(both_ok_position) - 1, -1, -1):
                        if both_ok_position[j]['section'] == self.static_section:
                            both_ok_position.pop(j)
                    for i in range(hours):
                        if len(both_ok_position) == 0:
                            return False
                        # 随机选择可排课程位置：
                        time = random.randint(0, len(both_ok_position) - 1)
                        day = both_ok_position[time]['week'] - 1
                        section = both_ok_position[time]['section'] - 1
                        # Assign class to schedule:
                        self.schedules[cla][day][section] = subj
                        self.schedules[cla][day][section + 1] = subj
                        self.schedules[cla_lock][day][section] = subj_lock
                        self.schedules[cla_lock][day][section + 1] = subj_lock
                        self.day_classes[cla][day][subj] = 1
                        self.day_classes[cla_lock][day][subj] = 1
                        # 排课后，约束条件增加位置2： remove adjacent day position
                        self.remove_linked_next_day(subj, cla, both_ok_position, day + 1)
                        # after assign update all available linked course time in cla:
                        # day and section - index
                        for k in range(len(cla_time_linked) - 1, -1, -1):
                            value = cla_time_linked[k]
                            if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                                cla_time_linked.pop(k)
                        for k in range(len(cla_time_locked_linked) - 1, -1, -1):
                            value = cla_time_locked_linked[k]
                            if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                                cla_time_locked_linked.pop(k)
                        for j in range(len(both_ok_position) - 1, -1, -1):
                            value = both_ok_position[j]
                            if value['week'] - 1 == day:
                                both_ok_position.pop(j)
                    # Update class:subj status:
                    self.classes_subj[cla][subj] = 1
                    self.classes_subj[cla_lock][subj_lock] = 1
        return True

    def assign_locked_course_free(self):
        self.get_all_free_position()
        self.update_all_course_hours()
        # init status: class_subj = 0
        # bug 6: not all equals to 0, manual schedule should stay 1:
        for i in range(len(self.classes_subj)):
            for subj in self.classes_subj[i]:
                self.classes_subj[i][subj] = 0
        for cla in range(self.class_amount):
            free_position_cla = self.free_position_all_classes[cla]
            for subj in self.class_course_teacher_info[cla + 1]:
                key = str(self.classes[cla]) + '-' + subj
                related_cla = self.get_related_class(cla, subj)
                if key in self.locked_cla_course:
                    # cla_lock and cla should be index
                    cla_lock = self.class_num[self.locked_cla_course[key]['class']] - 1
                    subj_lock = self.locked_cla_course[key]['subject']
                    free_position_cla_lock = self.free_position_all_classes[cla_lock]
                    related_cla_lock = self.get_related_class(cla_lock, subj_lock)
                    if self.free_course_hours[cla][subj] == 0 or self.free_course_hours[cla_lock][subj_lock] == 0:
                        continue
                    if self.classes_subj[cla][subj] == 1 or self.classes_subj[cla_lock][subj_lock] == 1:
                        continue
                    related_conflict_position = []
                    free_position_subj = copy.deepcopy(self.free_position_all_classes[cla])
                    free_position_subj_lock = copy.deepcopy(self.free_position_all_classes[cla_lock])
                    self.remove_subj_no_assign_free(subj, free_position_subj)
                    self.remove_subj_no_assign_free(subj_lock, free_position_subj_lock)
                    self.remove_teacher_no_assign_free(subj, cla, free_position_subj)
                    self.remove_teacher_no_assign_free(subj_lock, cla_lock, free_position_subj_lock)
                    # 挖去关联班级已排冲突的位置
                    self.remove_already_related_conflict_course_position(subj, related_cla, free_position_subj)
                    self.remove_already_related_conflict_course_position(subj_lock, related_cla_lock,
                                                                         free_position_subj_lock)
                    # add condition : bug 9 position_subj 挖去当天已排
                    for l in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[l]
                        if self.day_classes[cla][value['week']][subj] == 1:
                            free_position_subj.pop(l)
                    for l in range(len(free_position_subj_lock) - 1, -1, -1):
                        value = free_position_subj_lock[l]
                        if self.day_classes[cla][value['week']][subj] == 1:
                            free_position_subj_lock.pop(l)
                    """
                        连锁班级可排位置 然后取交集
                    """
                    both_ok_position = []
                    for grid in free_position_subj:
                        if grid in free_position_subj_lock:
                            both_ok_position.append(grid)
                    hours = min(self.free_course_hours[cla][subj], self.free_course_hours[cla_lock][subj_lock])
                    for hour in range(hours):
                        # if conflict: swap with already assigned
                        if len(both_ok_position) == 0:
                            print("No locked course position!")
                            return False
                        # assign to free position:
                        time = random.randint(0, len(both_ok_position) - 1)
                        day = both_ok_position[time]['week']
                        section = both_ok_position[time]['section']
                        related_conflict_position.append([day, section])
                        # assign subj to cla schedule:
                        self.schedules[cla][day][section] = subj
                        self.schedules[cla_lock][day][section] = subj_lock
                        # after assign update all available free course time in cla:
                        for k in range(len(free_position_cla) - 1, -1, -1):
                            value = free_position_cla[k]
                            if value['week'] == day and value['section'] == section:
                                free_position_cla.pop(k)
                                break
                        for k in range(len(free_position_cla_lock) - 1, -1, -1):
                            value = free_position_cla_lock[k]
                            if value['week'] == day and value['section'] == section:
                                free_position_cla_lock.pop(k)
                                break
                        for j in range(len(both_ok_position) - 1, -1, -1):
                            value = both_ok_position[j]
                            if value['week'] == day:
                                both_ok_position.pop(j)
                        # update status:
                        self.day_classes[cla][day][subj] = 1
                        self.day_classes[cla_lock][day][subj_lock] = 1
                    self.classes_subj[cla][subj] = 1
                    self.classes_subj[cla_lock][subj_lock] = 1
        return True

    """
    # 03: Assign linked courses:
        安排连堂课：
        temp_cla_linked_times: 记录班级可排连堂课的时间，每次排完课后update
        temp_subj_linked_times: 记录科目可排连堂课的时间，初始化：周可排（已-固定位置不排）- 固定科目不排 - 固定教师不排
                                排课后，temp_subj_linked_times - 当天已排 - 相邻天位置（remove_linked_next_day）
        安排顺序：
            遍历班级：
                遍历科目：
                    遍历科目课时：
                        ...
                安排关联班级同一科目：
    """

    def assign_linked_courses(self):
        self.update_linked_positions()
        for cla in range(self.class_amount):
            # update linked_course_time_all_classes[cla]
            temp_cla_linked_times = self.linked_course_time_all_classes[cla]
            math_linked_day = []
            for subj in self.class_course_teacher_info[cla + 1]:
                related_cla = self.get_related_class(cla, subj)
                # Already assigned:
                if self.classes_subj[cla][subj] == 1:
                    continue
                linked_course_num = self.linked_course_hours_classes[cla][subj]
                # No linked_hours left:
                if linked_course_num == 0:
                    continue
                # Assign class 1 subj from "语":
                temp_subj_linked_times = copy.deepcopy(temp_cla_linked_times)
                # 约束条件增加位置1：
                # linked_course subj_time - subj_no_assign
                self.remove_subj_no_assign(subj, temp_subj_linked_times)
                # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
                self.remove_already_related_conflict_course_position_linked(subj, related_cla, temp_subj_linked_times)
                # linked_course subj_time - teacher_no_assign
                self.remove_teacher_no_assign(subj, cla, temp_subj_linked_times)
                # 移除所有已经排了2天连堂课的天数：
                self.remove_already_assigned_two_linked_classes(cla, temp_subj_linked_times)
                related_conflict_position = []
                for i in range(linked_course_num):
                    if len(temp_subj_linked_times) == 0:
                        print("Start swap linked course position!")
                        if self.swap_conflict_linked(temp_cla_linked_times, cla, subj, temp_subj_linked_times):
                            continue
                        else:
                            print("No linked course swap position!")
                            return False
                    # 随机选择可排课程位置：
                    time = random.randint(0, len(temp_subj_linked_times) - 1)
                    day = temp_subj_linked_times[time]['week'] - 1
                    section = temp_subj_linked_times[time]['section'] - 1
                    # add to related conflict pos:
                    related_conflict_position.append([day, section - 1])
                    related_conflict_position.append([day, section])
                    related_conflict_position.append([day, section + 1])
                    # assign subj to cla schedule: \ section and day are index
                    self.schedules[cla][day][section] = subj
                    self.schedules[cla][day][section + 1] = subj
                    # add to math_linked_day
                    if subj == '数':
                        math_linked_day.append(day + 1)
                    # format: real date
                    self.already_assigned_linked_positions[cla].append({'week': day + 1, 'section': section + 1})
                    # update class:day:subj status:
                    self.day_classes[cla][day][subj] = 1
                    # 排课后，约束条件增加位置2： remove adjacent day position
                    self.remove_linked_next_day(subj, cla, temp_subj_linked_times, day + 1)
                    # after assign update all available linked course time in cla:
                    # day and section - index
                    for k in range(len(temp_cla_linked_times) - 1, -1, -1):
                        value = temp_cla_linked_times[k]
                        if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                                (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                                (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                            temp_cla_linked_times.pop(k)
                    for j in range(len(temp_subj_linked_times) - 1, -1, -1):
                        value = temp_subj_linked_times[j]
                        if value['week'] - 1 == day:
                            temp_subj_linked_times.pop(j)
                # Update class:subj status:
                self.classes_subj[cla][subj] = 1
                # Assign subj in related class(current cla):
                if not self.assign_related_class(related_cla, subj, related_conflict_position, math_linked_day):
                    return False
        return True

    # 03 - 00: update linked course positions: 已经排好课位置，以及排好课前面的位置都不能排连堂课
    def update_linked_positions(self):
        for cla in range(self.class_amount):
            i = len(self.linked_course_time_all_classes[cla]) - 1
            while i >= 0:
                day = self.linked_course_time_all_classes[cla][i]['week'] - 1
                section = self.linked_course_time_all_classes[cla][i]['section'] - 1
                # 没有存储最后一个section，所以如果最后一个有课，sec 6 不会被删除 -- bug 18
                if self.schedules[cla][day][section] is not None:
                    self.linked_course_time_all_classes[cla].pop(i)
                    if i - 1 >= 0 and section != 2:
                        self.linked_course_time_all_classes[cla].pop(i - 1)
                        i -= 1
                elif self.schedules[cla][day][section] is None and \
                        section + 1 < self.static_week and \
                        self.schedules[cla][day][section + 1] is not None:
                    self.linked_course_time_all_classes[cla].pop(i)
                # remove section == 7
                elif section == self.static_section - 1:
                    self.linked_course_time_all_classes[cla].pop(i)
                i -= 1

    # 03 - 01: find related class:
    # return index:
    def get_related_class(self, cla, subj):
        related_cla = []
        # bug: cla + 1
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(self.class_course_teacher_info)):
            if subj not in self.class_course_teacher_info[i + 1]:
                continue
            # bug2: i, cla are indexes
            if i != cla and self.class_course_teacher_info[i + 1][subj] == teacher:
                related_cla.append(i)
        return related_cla

    # 03 - 02: assign related class，subj：
    def assign_related_class(self, related_cla, subj, related_conflict_position, math_linked_day):
        for cla in related_cla:
            # check if subj already assigned:
            if self.classes_subj[cla][subj] == 1:
                continue
            linked_course_num = self.linked_course_hours_classes[cla][subj]
            temp_cla_linked_times = self.linked_course_time_all_classes[cla]
            temp_subj_linked_times = copy.deepcopy(self.linked_course_time_all_classes[cla])
            # 约束条件：
            # subj for cla need to remove all related position:
            self.remove_related_position(temp_subj_linked_times, related_conflict_position)
            # remove no assign class positions:
            self.remove_subj_no_assign(subj, temp_subj_linked_times)
            self.remove_teacher_no_assign(subj, cla, temp_subj_linked_times)
            # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
            related_cla_cur = self.get_related_class(cla, subj)
            self.remove_already_related_conflict_course_position_linked(subj, related_cla_cur, temp_subj_linked_times)
            # 移除所有已经排了2天连堂课的天数：
            self.remove_already_assigned_two_linked_classes(cla, temp_subj_linked_times)
            for i in range(linked_course_num):
                # Math linked course should be assigned on the same day
                if subj == '数':
                    if not self.remove_other_day(subj, cla, temp_subj_linked_times, math_linked_day):
                        return False
                if len(temp_subj_linked_times) == 0:
                    print("Start swap linked course")
                    if self.swap_conflict_linked(temp_cla_linked_times, cla, subj, temp_subj_linked_times):
                        continue
                    else:
                        print("No linked course swap position!")
                        return False
                time = random.randint(0, len(temp_subj_linked_times) - 1)
                day = temp_subj_linked_times[time]['week'] - 1
                section = temp_subj_linked_times[time]['section'] - 1
                related_conflict_position.append([day, section])
                related_conflict_position.append([day, section + 1])
                # assign subj to cla schedule:
                self.schedules[cla][day][section] = subj
                self.schedules[cla][day][section + 1] = subj
                # remove adjacent day position:
                self.remove_linked_next_day(subj, cla, temp_subj_linked_times, day + 1)
                # update class:day:subj status:
                self.day_classes[cla][day][subj] = 1
                # after assign update all available linked course time in cla:
                for k in range(len(temp_cla_linked_times) - 1, -1, -1):
                    value = temp_cla_linked_times[k]
                    if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                            (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                            (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                        temp_cla_linked_times.pop(k)
                for j in range(len(temp_subj_linked_times) - 1, -1, -1):
                    value = temp_subj_linked_times[j]
                    if value['week'] - 1 == day:
                        temp_subj_linked_times.pop(j)
            # Update class:subj status:
            self.classes_subj[cla][subj] = 1
        return True

    # 03 - 02 - 01: remove related conflict position:
    def remove_related_position(self, temp_subj_linked_times, related_conflict_position):
        for i in range(len(temp_subj_linked_times) - 1, -1, -1):
            for j in range(len(related_conflict_position)):
                if temp_subj_linked_times[i]['week'] - 1 == related_conflict_position[j][0] and \
                        temp_subj_linked_times[i]['section'] - 1 == related_conflict_position[j][1]:
                    temp_subj_linked_times.pop(i)
                    break

    # add condition function:
    # 03 - 03 : subj_time - subj_no_assign_time / here free time is not index
    def remove_subj_no_assign(self, subj, subj_time):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if subj in self.subj_no_assign:
                for no_assign_time in self.subj_no_assign[subj]:
                    # bug 12: linked course need to consider double position:
                    # if time['week'] == no_assign_time['week'] and time['section'] == no_assign_time['section']:
                    if (time['week'] == no_assign_time['week'] and time['section'] == no_assign_time['section']) or \
                            (time['week'] == no_assign_time['week'] and time['section'] + 1 == no_assign_time[
                                'section']):
                        subj_time.pop(i)
                        break
            else:
                break

    # 03 - 04 : subj_time - teacher_no_assign_time
    def remove_teacher_no_assign(self, subj, cla, subj_time):
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if teacher in self.teacher_no_assign:
                for no_assign_time in self.teacher_no_assign[teacher]:
                    if (time['week'] == no_assign_time['week'] and time['section'] == no_assign_time['section']) or \
                            (time['week'] == no_assign_time['week'] and time['section'] + 1 == no_assign_time[
                                'section']):
                        subj_time.pop(i)
                        break

    # 03 - 05 : remove next linked position : day is not index
    def remove_linked_next_day(self, subj, cla, subj_time, day):
        next = day + 1
        pre = day - 1
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if next <= self.static_week:
                if time['week'] == next:
                    subj_time.pop(i)
                    continue
            if pre >= 1:
                if time['week'] == pre:
                    subj_time.pop(i)

    # 03 - 06 : SWAP LINKED CONFLICT POINT - All date format are real date
    # 没考虑同一天连堂课交换
    # 空余位置与已排连堂课交换，free position 排除一天已排2连堂情况
    def swap_conflict_linked(self, free_position_cla, cla, subj, free_position_subj):
        # 如果当天连堂课 >=2 remove 这个free_position_cla 只移除已经安排的课程
        # self.remove_already_assigned_two_linked_classes(cla, free_position_cla)
        for i in range(len(free_position_cla)):
            # current position b
            position = free_position_cla[i]
            day_b = position['week']
            section_b = position['section']
            teacher_b = self.class_course_teacher_info[cla + 1][subj]
            subj_b = subj
            # 如果连堂课数量 >= 2 跳过 选取下一个位置
            # if already assigned course >= 2
            count = 0
            for j in range(self.static_section):
                if self.schedules[cla][day_b - 1][j] is not None and (j + 1 < self.static_section) and \
                        self.schedules[cla][day_b - 1][j] == self.schedules[cla][day_b - 1][j + 1]:
                    if self.schedules[cla][day_b - 1][j] != '走':
                        count += 1
            # 提取公共参数： 最大连堂课个数
            if count >= self.link_course_count:
                continue
            for j in range(len(self.already_assigned_linked_positions[cla])):
                # if already assigned position day already has subj:
                already_position = self.already_assigned_linked_positions[cla]
                if self.day_classes[cla][already_position[j]['week'] - 1][subj] == 1:
                    continue
                # get position to swap a:
                day_a = self.already_assigned_linked_positions[cla][j]['week']
                section_a = self.already_assigned_linked_positions[cla][j]['section']
                subj_a = self.schedules[cla][day_a - 1][section_a - 1]
                teacher_a = self.class_course_teacher_info[cla + 1][subj_a]
                # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
                # b: current free position a: already assigned position
                same_day = False
                if day_a == day_b:
                    same_day = True
                if not same_day and self.day_classes[cla][day_b - 1][subj_a] == 1 or self.day_classes[cla][day_a - 1][
                    subj_b] == 1:
                    continue
                # judge whether could swap:
                # if put b to a: check related class position b teacher is not teacher_b
                flag = True
                # bug 10: re_cla get teacher should plus 1
                # Get related class:
                relate_cla_a = self.get_related_class(cla, subj_a)
                relate_cla_b = self.get_related_class(cla, subj_b)
                for re_cla in relate_cla_b:
                    if (self.schedules[re_cla][day_a - 1][section_a - 1] == subj_b) or \
                            (self.schedules[re_cla][day_a - 1][section_a] == subj_b):
                        flag = False
                        break
                for re_cla in relate_cla_a:
                    # if put a to b:
                    if (self.schedules[re_cla][day_b - 1][section_b - 1] == subj_a) or \
                            (self.schedules[re_cla][day_b - 1][section_b] == subj_a):
                        flag = False
                        break
                if not flag:
                    continue
                # if a not in subj_no_assign position: day_a,section_a is not in subj_no_assign(subj_b)
                for position in self.subj_no_assign[subj_b]:
                    if (position['week'] == day_a and position['section'] == section_a) or \
                            (position['week'] == day_a and position['section'] == section_a + 1):
                        flag = False
                        break
                if not flag:
                    continue
                # if b not in subj_no_assign position: day_b,section_b is not in subj_no_assign(subj_a)
                for position in self.subj_no_assign[subj_a]:
                    if position['week'] == day_b and position['section'] == section_b or \
                            position['week'] == day_b and position['section'] == section_b + 1:
                        flag = False
                        break
                if not flag:
                    continue
                # after swap teacher_a not in teacher_no_assign:
                if teacher_a in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_a]:
                        if time['week'] == day_b and time['section'] == section_b or \
                                time['week'] == day_b and time['section'] == section_b + 1:
                            flag = False
                            break
                if not flag:
                    continue
                # after swap teacher_b not in teacher_no_assign:
                if teacher_b in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_b]:
                        if time['week'] == day_a and time['section'] == section_a or \
                                time['week'] == day_a and time['section'] == section_a + 1:
                            flag = False
                            break
                # After swap, no next_day with same course:
                pre_day_a = max(0, day_a - 1)
                next_day_a = min(self.static_week, day_a + 1)
                pre_day_b = max(0, day_b - 1)
                next_day_b = min(self.static_week, day_b + 1)
                # bug 15: 应该周边不能有连堂课 不是 不能有该科目
                if self.day_classes[cla][pre_day_a - 1][subj_b] == 1 or \
                        self.day_classes[cla][next_day_a - 1][subj_b] == 1:
                    flag = False
                if self.day_classes[cla][pre_day_b - 1][subj_a] == 1 or \
                        self.day_classes[cla][next_day_b - 1][subj_a] == 1:
                    flag = False
                if flag:
                    # swap in schedules and update status: a and b in cla
                    self.schedules[cla][day_b - 1][section_b - 1] = subj_a
                    self.schedules[cla][day_b - 1][section_b] = subj_a
                    self.schedules[cla][day_a - 1][section_a - 1] = subj_b
                    self.schedules[cla][day_a - 1][section_a] = subj_b
                    if not same_day:
                        self.day_classes[cla][day_b - 1][subj_a] = 1
                        self.day_classes[cla][day_b - 1][subj_b] = 0
                        self.day_classes[cla][day_a - 1][subj_b] = 1
                        self.day_classes[cla][day_a - 1][subj_a] = 0
                    # bug5: update already assigned:
                    self.already_assigned_linked_positions[cla].append({'week': day_b, 'section': section_b})
                    # update free position_cla and free position_subj:
                    for k in range(len(free_position_cla) - 1, -1, -1):
                        value = free_position_cla[k]
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_cla.pop(k)
                    for j in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[j]
                        if value['week'] == day_a:
                            free_position_subj.pop(j)
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_subj.pop(j)
                    return True
        return False

    # 03 - 07 : subj_time - subj_already_linked_course_position ( including pre and cur)
    def remove_already_related_conflict_course_position_linked(self, subj, related_cla, subj_time):
        # bug 15 迭代器
        for i in range(len(subj_time) - 1, -1, -1):
            for re_cla in related_cla:
                # try:
                week = subj_time[i]['week'] - 1
                section = subj_time[i]['section'] - 1
                # remove cur and pre
                if self.schedules[re_cla][week][section] == subj:
                    subj_time.pop(i)
                    break
                if section + 1 < self.static_section and \
                        self.schedules[re_cla][week][section + 1] == subj:
                    subj_time.pop(i)
                    break
                # except IndexError as err:
                #     print("Index Exception")

    # 03 - 08 : remove all other positions : day is not index
    def remove_other_day(self, subj, cla, subj_time, math_linked_day):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if time['week'] not in math_linked_day:
                subj_time.pop(i)
        if len(subj_time) == 0:
            return False
        return True

    # 03 - 09 : 移除所有已经排了2天连堂课的天数：
    def remove_already_assigned_two_linked_classes(self, cla, subj_time):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            # if already assigned course >= 2
            count = 0
            for j in range(self.static_section):
                if self.schedules[cla][time['week'] - 1][j] is not None and (j + 1 < self.static_section) and \
                        self.schedules[cla][time['week'] - 1][j] == self.schedules[cla][time['week'] - 1][j + 1]:
                    if self.schedules[cla][time['week'] - 1][j] != '走':
                        count += 1
            # 提取公共参数： 最大连堂课个数
            if count >= self.link_course_count:
                subj_time.pop(i)

    # 04: Assign free courses:
    # from here: day, section presents index in schedule
    def assign_free_courses(self):
        self.get_all_free_position()
        self.update_all_course_hours()
        # init status: class_subj = 0
        # bug 6: not all equals to 0, manual schedule should stay 1:
        for i in range(len(self.classes_subj)):
            for subj in self.classes_subj[i]:
                self.classes_subj[i][subj] = 0
        for cla in range(self.class_amount):
            free_position_cla = self.free_position_all_classes[cla]
            for subj in self.class_course_teacher_info[cla + 1]:
                related_cla = self.get_related_class(cla, subj)
                if self.classes_subj[cla][subj] == 1:
                    continue
                related_conflict_position = []
                free_position_subj = copy.deepcopy(self.free_position_all_classes[cla])
                self.remove_subj_no_assign_free(subj, free_position_subj)
                self.remove_teacher_no_assign_free(subj, cla, free_position_subj)
                # 挖去关联班级已排冲突的位置
                self.remove_already_related_conflict_course_position(subj, related_cla, free_position_subj)
                # add condition : bug 9 position_subj 挖去当天已排
                for l in range(len(free_position_subj) - 1, -1, -1):
                    value = free_position_subj[l]
                    if self.day_classes[cla][value['week']][subj] == 1:
                        free_position_subj.pop(l)
                if self.free_course_hours[cla][subj] == 0:
                    continue
                for hours in range(self.free_course_hours[cla][subj]):
                    # if conflict: swap with already assigned
                    if len(free_position_subj) == 0:
                        print("Start swap")
                        if self.swap_conflict(free_position_cla, cla, subj, free_position_subj):
                            continue
                        else:
                            # 尝试交换以排好的两个点，多次判断是否可以交换
                            if not self.random_swap_already_assigned(free_position_cla, cla, subj, free_position_subj):
                                print("No swap position!")
                                self.conflict_course_list.append({cla: subj})
                                continue
                            else:
                                continue
                            # self.conflict_course_list.append({cla: subj})
                            # continue
                    # assign to free position:
                    time = random.randint(0, len(free_position_subj) - 1)
                    day = free_position_subj[time]['week']
                    section = free_position_subj[time]['section']
                    related_conflict_position.append([day, section])
                    # assign subj to cla schedule:
                    self.schedules[cla][day][section] = subj
                    self.already_assigned_free_positions[cla].append({'week': day, 'section': section})
                    # after assign update all available free course time in cla:
                    for k in range(len(free_position_cla) - 1, -1, -1):
                        value = free_position_cla[k]
                        if value['week'] == day and value['section'] == section:
                            free_position_cla.pop(k)
                            break
                    for j in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[j]
                        if value['week'] == day:
                            free_position_subj.pop(j)
                    # update status:
                    self.day_classes[cla][day][subj] = 1
                self.classes_subj[cla][subj] = 1
                # Assign subj in related class(current cla):
                if not self.assign_related_class_free(related_cla, subj, related_conflict_position):
                    return False
        return True

    # 04 - 01 :  get all free position
    def get_all_free_position(self):
        self.free_position_all_classes = []
        for cla in range(self.class_amount):
            free_position_a_class = []
            for day in range(len(self.schedules[cla])):
                for section in range(len(self.schedules[cla][day])):
                    if self.schedules[cla][day][section] is None:
                        time = {'week': day, 'section': section}
                        free_position_a_class.append(time)
            self.free_position_all_classes.append(free_position_a_class)

    # 04 - 02 : update all course hours:
    def update_all_course_hours(self):
        self.free_course_hours = []
        for cla in range(self.class_amount):
            dic = {}
            for subj in self.class_course_teacher_info[cla + 1]:
                dic[subj] = self.course_hours[subj]
                # 遍历课表，出现该科目 dic[subj] - 1
                for day in range(self.static_week):
                    for section in range(self.static_section):
                        if self.schedules[cla][day][section] == subj:
                            dic[subj] -= 1
                # 额外安排课程的free_course_hours = 2:
                for unit in self.extra_assign_time[self.classes[cla]]:
                    if unit['course'] == subj:
                        dic[subj] = unit['num']
            self.free_course_hours.append(dic)
        # for manual assigned subj in specific cla set to 0 :
        # for course in self.manual_assign_time:
        #     self.free_course_hours[self.class_num[course['class']] - 1][course['course']] = 0

    # 04 - 03: assign related free classes:
    def assign_related_class_free(self, related_cla, subj, related_conflict_position):
        for cla in related_cla:
            # check if subj already assigned:
            if self.classes_subj[cla][subj] == 1:
                continue
            related_class = self.get_related_class(cla, subj)
            free_course_num = self.free_course_hours[cla][subj]
            temp_cla_free_times = self.free_position_all_classes[cla]
            temp_subj_free_times = copy.deepcopy(self.free_position_all_classes[cla])
            # subj for cla need to remove all related position:
            self.remove_related_position_free(temp_subj_free_times, related_conflict_position)
            self.remove_subj_no_assign_free(subj, temp_subj_free_times)
            self.remove_teacher_no_assign_free(subj, cla, temp_subj_free_times)
            # 挖去关联班级已排的该课程位置
            self.remove_already_related_conflict_course_position(subj, related_class, temp_subj_free_times)
            # add condition : subj_time 挖去当天已排
            for l in range(len(temp_subj_free_times) - 1, -1, -1):
                value = temp_subj_free_times[l]
                if self.day_classes[cla][value['week']][subj] == 1:
                    temp_subj_free_times.pop(l)
            for i in range(free_course_num):
                # if conflict:
                if len(temp_subj_free_times) == 0:
                    print("No related position left!")
                    if self.swap_conflict(temp_cla_free_times, cla, subj, temp_subj_free_times):
                        continue
                    else:
                        # 尝试交换以排好的两个点，多次判断是否可以交换
                        if not self.random_swap_already_assigned(temp_cla_free_times, cla, subj, temp_subj_free_times):
                            print("No swap position!")
                            self.conflict_course_list.append({cla: subj})
                            continue
                        else:
                            continue
                        # self.conflict_course_list.append({cla: subj})
                        # continue
                time = random.randint(0, len(temp_subj_free_times) - 1)
                day = temp_subj_free_times[time]['week']
                section = temp_subj_free_times[time]['section']
                related_conflict_position.append([day, section])
                # assign subj to cla schedule:
                self.schedules[cla][day][section] = subj
                # update class:day:subj status:
                self.day_classes[cla][day][subj] = 1
                # update already assigned status:
                self.already_assigned_free_positions[cla].append({'week': day, 'section': section})
                # after assign update all available free course time in cla:
                for k in range(len(temp_cla_free_times) - 1, -1, -1):
                    value = temp_cla_free_times[k]
                    if value['week'] == day and value['section'] == section:
                        temp_cla_free_times.pop(k)
                for j in range(len(temp_subj_free_times) - 1, -1, -1):
                    value = temp_subj_free_times[j]
                    if value['week'] == day:
                        temp_subj_free_times.pop(j)
            # Update class:subj status:
            self.classes_subj[cla][subj] = 1
        return True

    # 04 - 04 : SWAP CONFLICT POINT
    # 如果与当天课程交换，不能是同一门
    # 如果当天已有该课程， 不能交换， continue
    def swap_conflict(self, free_position_cla, cla, subj, free_position_subj):
        for i in range(len(free_position_cla)):
            # current position b
            position = free_position_cla[i]
            day_b = position['week']
            section_b = position['section']
            teacher_b = self.class_course_teacher_info[cla + 1][subj]
            subj_b = subj
            for j in range(len(self.already_assigned_free_positions[cla])):
                # if already assigned position day already has subj:
                already_position = self.already_assigned_free_positions[cla]
                if self.day_classes[cla][already_position[j]['week']][subj] == 1:
                    continue
                # get position to swap a:
                day_a = self.already_assigned_free_positions[cla][j]['week']
                section_a = self.already_assigned_free_positions[cla][j]['section']
                subj_a = self.schedules[cla][day_a][section_a]
                teacher_a = self.class_course_teacher_info[cla + 1][subj_a]
                # if free_position == already_assigend position, continue
                if day_a == day_b and section_a == section_b:
                    continue
                # bug 19: 如果空余位置当天已经存在该课程，该位置无效
                # bug 20 free_position_cla 空余位置排除已经有了该课程的
                if self.day_classes[cla][day_a][subj_b] == 1 or self.day_classes[cla][day_b][subj_b] == 1:
                    continue
                if subj_a == subj_b:
                    continue
                # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
                same_day = False
                # bug 14： not consider day_a == day_b
                # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
                if day_a != day_b and (
                        self.day_classes[cla][day_b][subj_a] == 1 or self.day_classes[cla][day_a][subj_b] == 1):
                    continue
                if day_a == day_b:
                    same_day = True
                # judge whether could swap:
                # if put b to a: check related class position b teacher is not teacher_b
                flag = True
                # bug 10: re_cla get teacher should plus 1
                # bug 13: 同一个班级下，不科目的关联班级不一样
                # related class:
                relate_cla_b = self.get_related_class(cla, subj_b)
                relate_cla_a = self.get_related_class(cla, subj_a)
                for re_cla in relate_cla_b:
                    if self.schedules[re_cla][day_a][section_a] == subj_b:
                        flag = False
                        break
                for re_cla in relate_cla_a:
                    # if put a to b:
                    if self.schedules[re_cla][day_b][section_b] == subj_a:
                        flag = False
                        break
                if not flag:
                    continue
                # if a not in subj_no_assign position: day_a,section_a is not in subj_no_assign(subj_b)
                for position in self.subj_no_assign[subj_b]:
                    if position['week'] - 1 == day_a and position['section'] - 1 == section_a:
                        flag = False
                        break
                if not flag:
                    continue
                # if b not in subj_no_assign position: day_b,section_b is not in subj_no_assign(subj_a)
                for position in self.subj_no_assign[subj_a]:
                    if position['week'] - 1 == day_b and position['section'] - 1 == section_b:
                        flag = False
                        break
                if not flag:
                    continue
                # after swap teacher_a not in teacher_no_assign:
                if teacher_a in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_a]:
                        if time['week'] - 1 == day_b and time['section'] - 1 == section_b:
                            flag = False
                            break
                if not flag:
                    continue
                # after swap teacher_b not in teacher_no_assign:
                if teacher_b in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_b]:
                        if time['week'] - 1 == day_a and time['section'] - 1 == section_a:
                            flag = False
                            break
                if flag:
                    # swap in schedules and update status: a and b in cla
                    self.schedules[cla][day_b][section_b] = subj_a
                    self.schedules[cla][day_a][section_a] = subj_b
                    if not same_day:
                        self.day_classes[cla][day_b][subj_a] = 1
                        self.day_classes[cla][day_b][subj_b] = 0
                        self.day_classes[cla][day_a][subj_b] = 1
                        self.day_classes[cla][day_a][subj_a] = 0
                    # bug5: update already assigned:
                    self.already_assigned_free_positions[cla].append({'week': day_b, 'section': section_b})
                    # update free position_cla and free position_subj:
                    for k in range(len(free_position_cla) - 1, -1, -1):
                        value = free_position_cla[k]
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_cla.pop(k)
                    for j in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[j]
                        if value['week'] == day_a:
                            free_position_subj.pop(j)
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_subj.pop(j)
                    return True
        return False

    # 04 - 05: remove related free conflict position:
    # bug 7: 2 system day and week --> index
    def remove_related_position_free(self, temp_subj_free_times, related_conflict_position):
        for i in range(len(temp_subj_free_times) - 1, -1, -1):
            for j in range(len(related_conflict_position)):
                if temp_subj_free_times[i]['week'] == related_conflict_position[j][0] and \
                        temp_subj_free_times[i]['section'] == related_conflict_position[j][1]:
                    temp_subj_free_times.pop(i)
                    break

    # 04 - 06 : subj_time - subj_no_assign_time / here free time is not index
    def remove_subj_no_assign_free(self, subj, subj_time):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if subj in self.subj_no_assign:
                for no_assign_time in self.subj_no_assign[subj]:
                    if time['week'] == no_assign_time['week'] - 1 and time['section'] == no_assign_time['section'] - 1:
                        subj_time.pop(i)

    # 04 - 07 : subj_time - teacher_no_assign_time
    def remove_teacher_no_assign_free(self, subj, cla, subj_time):
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if teacher in self.teacher_no_assign:
                for no_assign_time in self.teacher_no_assign[teacher]:
                    if time['week'] == no_assign_time['week'] - 1 and time['section'] == no_assign_time['section'] - 1:
                        subj_time.pop(i)

    # 04 - 08 : subj_time - subj_already_linked_course_position
    def remove_already_related_conflict_course_position(self, subj, related_cla, subj_time):
        for i in range(len(subj_time) - 1, -1, -1):
            for re_cla in related_cla:
                if self.schedules[re_cla][subj_time[i]['week']][subj_time[i]['section']] == subj:
                    subj_time.pop(i)
                    break

    # 04 - 09: swap 2 already assigned course to check if could swap conflict after
    def random_swap_already_assigned(self, free_position_cla, cla, subj, free_position_subj):
        # check if swap 2 positions whether could get a result:
        for i in range(len(self.already_assigned_free_positions[cla]) - 1):
            for j in range(i + 1, len(self.already_assigned_free_positions[cla])):
                # swap_two_positions:
                position_i = self.already_assigned_free_positions[cla][i]
                day_a = position_i['week']
                section_a = position_i['section']
                subj_a = self.schedules[cla][day_a][section_a]
                position_j = self.already_assigned_free_positions[cla][j]
                day_b = position_j['week']
                section_b = position_j['section']
                subj_b = self.schedules[cla][day_b][section_b]
                same_day = False
                if day_b == day_a:
                    same_day = True
                if self.check_swap(cla, subj_a, subj_b, day_a, section_a, day_b, section_b):
                    # swap:
                    # swap in schedules and update status: a and b in cla
                    self.schedules[cla][day_b][section_b] = subj_a
                    self.schedules[cla][day_a][section_a] = subj_b
                    if not same_day:
                        self.day_classes[cla][day_b][subj_a] = 1
                        self.day_classes[cla][day_b][subj_b] = 0
                        self.day_classes[cla][day_a][subj_b] = 1
                        self.day_classes[cla][day_a][subj_a] = 0
                    # After swap check if could swap:
                    if self.swap_conflict(free_position_cla, cla, subj, free_position_subj):
                        return True
                    else:
                        # Swap back: =============================================================
                        self.schedules[cla][day_b][section_b] = subj_b
                        self.schedules[cla][day_a][section_a] = subj_a
                        if not same_day:
                            self.day_classes[cla][day_b][subj_a] = 0
                            self.day_classes[cla][day_b][subj_b] = 1
                            self.day_classes[cla][day_a][subj_b] = 0
                            self.day_classes[cla][day_a][subj_a] = 1
        return False

    # 04 - 10: 查看两点是否可交换 - bug 16 挖去当天已排位置 bug 19: 当天已存在该课程，不交换
    def check_swap(self, cla, subj_a, subj_b, day_a, section_a, day_b, section_b):
        # if free_position == already_assigend position, continue
        if day_a == day_b and section_a == section_b:
            return False
        # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
        same_day = False
        # bug 14： not consider day_a == day_b
        # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
        if day_a != day_b and (
                self.day_classes[cla][day_b][subj_a] == 1 or self.day_classes[cla][day_a][subj_b] == 1):
            return False
        # bug 19: else: 导致两个正常的点，被认为是同一天。交换后更新科目安排状态存在问题。
        if day_a == day_b:
            same_day = True
        # judge whether could swap:
        # if put b to a: check related class position b teacher is not teacher_b
        flag = True
        # bug 10: re_cla get teacher should plus 1
        # bug 13: 同一个班级下，不科目的关联班级不一样
        # related class:
        relate_cla_b = self.get_related_class(cla, subj_b)
        relate_cla_a = self.get_related_class(cla, subj_a)
        teacher_a = self.class_course_teacher_info[cla + 1][subj_a]
        teacher_b = self.class_course_teacher_info[cla + 1][subj_b]
        for re_cla in relate_cla_b:
            if self.schedules[re_cla][day_a][section_a] == subj_b:
                flag = False
                return False
        for re_cla in relate_cla_a:
            # if put a to b:
            if self.schedules[re_cla][day_b][section_b] == subj_a:
                flag = False
                return False
        # if a not in subj_no_assign position: day_a,section_a is not in subj_no_assign(subj_b)
        for position in self.subj_no_assign[subj_b]:
            if position['week'] - 1 == day_a and position['section'] - 1 == section_a:
                flag = False
                return False
        # if b not in subj_no_assign position: day_b,section_b is not in subj_no_assign(subj_a)
        for position in self.subj_no_assign[subj_a]:
            if position['week'] - 1 == day_b and position['section'] - 1 == section_b:
                flag = False
                return False
        # after swap teacher_a not in teacher_no_assign:
        if teacher_a in self.teacher_no_assign:
            for time in self.teacher_no_assign[teacher_a]:
                if time['week'] - 1 == day_b and time['section'] - 1 == section_b:
                    flag = False
                    return False
        # after swap teacher_b not in teacher_no_assign:
        if teacher_b in self.teacher_no_assign:
            for time in self.teacher_no_assign[teacher_b]:
                if time['week'] - 1 == day_a and time['section'] - 1 == section_a:
                    flag = False
                    return False
        return flag

    def output_schedules(self):
        # 将课表输出到excel中
        workbook = openpyxl.Workbook()
        for classes in self.classes_name:
            worksheet = workbook.create_sheet(classes)
            schedule = self.schedules[self.class_num[classes] - 1]
            print(classes)
            for x in range(self.static_week):
                for y in range(self.static_section):
                    course = schedule[x][y]
                    if course is None:
                        continue
                    else:
                        value = classes + '-' + course
                        if course in self.courses:
                            value += ('-' + self.data.CLASSES_COURSE_TEACHER[self.class_num[classes]][course])
                        worksheet.cell(y + 1, x + 1, value)
        workbook.save('schedule.xlsx')

    def log_execution_time(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            start = time.perf_counter()
            iteration = 1
            for i in range(iteration):
                result = func(*args, **kwargs)
            end = time.perf_counter()
            print('{} took average {} ms'.format(func.__name__, (end - start) * 1000 / iteration))
            return result

        return wrapper

    #主函数：
    @log_execution_time
    def run(self, data, solve_conflict):
        iteration = 0
        conflict = 0
        min_conflict = sys.maxsize
        print(min_conflict)
        result = None
        while iteration < 1000:
            gc.collect()
            self.__init__(data)
            print('iteration:{}'.format(iteration))
            # 01: Assign fixed flow classes： -1
            self.assign_fixed()
            # 02: Assign locked courses:
            if (not self.assign_locked_course_linked()) or (not self.assign_locked_course_free()):
                continue
            # 03: Assign linked classes
            if not self.assign_linked_courses():
                continue
            # 04: Assign reminded(free) classes:
            if not self.assign_free_courses():
                continue
            print(self.schedules)
            conflict = len(self.conflict_course_list)
            for i in range(len(self.conflict_course_list)):
                print(self.conflict_course_list[i], len(self.conflict_course_list))

            # 解决冲突
            # auther:ice
            conflict_list = []
            for a_conflict in self.conflict_course_list:
                for key, value in a_conflict.items():
                    hava_flag = False
                    for class_name in self.class_num:
                        if self.class_num[class_name] == key:
                            hava_flag = True
                            break
                    if hava_flag:
                        conflict_list.append({'class': class_name, 'course': value})
            solve_conflict_flag = False
            if solve_conflict is not None:
                solve_conflict_flag = solve_conflict(self.schedules, conflict_list)
            print('是否解决所有冲突', solve_conflict_flag)
            if conflict < min_conflict:
                min_conflict = conflict
                result = self.schedules
                # self.output_schedules()
            if min_conflict == 0 or solve_conflict_flag:
                break
            iteration += 1
        print("min conflict number:{}".format(min_conflict))
        return result

    # datasource 格式转换
    @staticmethod
    def translate_from(datasource: data_source.DataSource):
        data = datasource
        data.LOCKED_CLA_COURSE = datasource.BIND_CLASSES
        # classes num
        classes_num = {}
        for i in range(len(data.CLASSES)):
            classes_num[data.CLASSES[i]] = i + 1
        data.CLASSES_NUM = classes_num
        # classes_course_teacher
        keys = list(data.CLASSES_COURSE_TEACHER.keys())
        for key in keys:
            new_key = classes_num[key]
            data.CLASSES_COURSE_TEACHER[new_key] = data.CLASSES_COURSE_TEACHER[key]
            data.CLASSES_COURSE_TEACHER.pop(key)
        data.FLOW_SIGN = '走'
        data.NOT_ASSIGN_SIGN = '不排'
        data.CLASS_NUM = len(data.CLASSES)
        return data

"""
7-15 添加：数学老师的不同班连堂课在同一天
          连堂课一天max 3 限制
          2，3 大课间无连堂课
     待完成： 自动生成连锁课程
             扩大算法搜索范围          
"""
